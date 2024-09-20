# -*- codeing = utf-8 -*-
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配`
import urllib.request, urllib.error  # 制定URL，获取网页数据
import openpyxl  # 进行excel操作
import pandas as pd
from tqdm import tqdm
import os

language='en'

def main_zh():
    # 爬取中文网页
    OlympicsList = pd.read_csv("Olympics_event/olympic_games.csv")["Olympic Game"]


    for Olympics in tqdm(OlympicsList):
        print(Olympics)
        SportList = pd.read_csv("./Olympics_event/" + Olympics + "_events.csv")[
            "Event English Name"]
        # SportList=['table-tennis']
        for sport in SportList:
            print(sport)
            baseurl = "https://olympics.com/zh/olympic-games/" + Olympics + "/results/" + sport  # 要爬取的网页链接
            # 1.爬取网页
            datalist = getData(baseurl)
            # 如果没有目录新建目录
            if not os.path.exists("./Olympics-result-zh/" + Olympics):
                os.makedirs("./Olympics-result-zh/" + Olympics)
            if datalist:
                savepath = "./Olympics-result-zh/" + Olympics + "/" + sport + ".xlsx"  # 当前目录新建XLS，存储进去
                # print(datalist)
                # 2.保存数据
                saveData(datalist, savepath, sport,language='zh')

def main_en():
    # 爬取英语网页
    OlympicsList = pd.read_csv("./Olympics_event/olympic_games.csv")["Olympic Game"]
    # OlympicsList=['rio-2016']
    # OlympicsList = ['tokyo-2020']

    for Olympics in tqdm(OlympicsList):
        print(Olympics)
        SportList = pd.read_csv("./Olympics_event/" + Olympics + "_events.csv")[
            "Event English Name"]
        # SportList=['table-tennis']
        for sport in SportList:
            print(sport)
            baseurl = "https://olympics.com/en/olympic-games/" + Olympics + "/results/" + sport  # 要爬取的网页链接
            # 1.爬取网页
            datalist = getData(baseurl)
            # 如果没有目录新建目录
            if not os.path.exists("./Olympics-result-en/" + Olympics):
                os.makedirs("./Olympics-result-en/" + Olympics)
            if datalist:
                savepath = "./Olympics-result-en/" + Olympics + "/" + sport + ".xlsx"  # 当前目录新建XLS，存储进去
                # print(datalist)
                # 2.保存数据
                saveData(datalist, savepath, sport,language='en')


def get_single_athlete(ite, event, datalist):
    # 项目是单人运动
    # find_all('div', {'data-cy': "single-athlete-award-card"})

    for it in ite:
        data = []  # 保存所有信息
        # 项目
        data.append(event)
        # print(sport)
        # 奖牌
        findMedel = it.find('span', {'data-cy': "medal-additional"})
        # Medel = re.findall(findMedel, item)
        Medel = findMedel.text
        data.append(Medel)

        # 运动员链接和名字
        findAthleteName = it.find('a', {'data-cy': "name"})
        if findAthleteName:  # 有link 运动员
            AthleteName = findAthleteName.text.title()
            AthleteLink = findAthleteName['href']
        else:  # 无link 运动员
            AthleteName = it.find('span', {'data-cy': "name-no-link"}).text.title()
            AthleteLink = 'None'

        data.append(AthleteLink)
        data.append(AthleteName)

        # 国家名字
        findCountry = it.find('div', {'data-cy': "flag-with-label"})
        findCountry = findCountry.find_all('span')[-1]
        CountryID = findCountry['data-cy']
        Country = findCountry.text
        data.append(CountryID)  # 国家代码
        data.append(Country)  # 国家
        # print(Country)
        datalist.append(data)

    return datalist


def get_two_athletes(ite, event, datalist):
    # 项目是双人运动
    # find_all('div', {'data-cy': "two-athletes-award-card"})
    for it in ite:

        # 项目
        # data.append(sport)

        # 奖牌
        findMedel = it.find('span', {'data-cy': "medal-additional"})
        Medel = findMedel.text
        # data.append(Medel)

        # 运动员链接和名字
        div = it.find('div', {'data-cy': "athlete-1-name/athlete-2-name"})

        # 初始化变量
        athlete_names = []
        athlete_links = []

        # 遍历 div 的所有子元素
        for element in div.children:
            if element.name == 'a':  # 如果是 <a> 标签

                athlete_names.append(element.text.strip().title())
                athlete_links.append(element.get('href'))
            elif isinstance(element, str):  # 如果是文本节点
                names = element.strip().split('/')
                for name in names:
                    if name:  # 只处理非空文本
                        athlete_names.append(name.strip().title())
                        athlete_links.append(None)

        # 将名字和链接格式化为 'Name1 / Name2' 和 'Link1 / Link2'

        # formatted_names = ' / '.join(athlete_names)
        # formatted_links = ' / '.join(link if link else 'No link' for link in athlete_links)

        # 输出结果
        # print(f"Names: {formatted_names}")
        # print(f"Links: {formatted_links}")

        # Athletelist = findAthleteName.find_all('a', {'data-cy': "link"})
        # print(Athletelist)
        # TwoAthleteName = Athletelist[0].text.title() + " /" + Athletelist[1].text.title()
        # TwoAthleteLink = Athletelist[0]['href'] + " /" + Athletelist[1]['href']

        # data.append(formatted_links)
        # data.append(formatted_names)

        # 国家
        findCountry = it.find('div', {'data-cy': "flag-with-label"})
        findCountry = findCountry.find_all('span')[-1]
        CountryID = findCountry['data-cy']
        Country = findCountry.text

        for name, link in zip(athlete_names, athlete_links):
            data = []  # 保存所有信息
            data.append(event)
            data.append(Medel)
            data.append(link)
            data.append(name)
            data.append(CountryID)  # 国家代码
            data.append(Country)  # 国家
            datalist.append(data)
    return datalist


def get_nation_team(ite, event, datalist):
    # 项目是国家团体
    # find_all('div', {'data-cy': "team-award-card "})
    for it in ite:
        data = []  # 保存所有信息
        # 项目
        data.append(event)

        # 奖牌
        findMedel = it.find('span', {'data-cy': "medal-additional"})
        # Medel = re.findall(findMedel, item)
        Medel = findMedel.text
        data.append(Medel)

        # 运动员链接和名字
        data.append(None)  # Link
        if language == 'zh':
            data.append("国家队")  # Name：国家队
        else:
            data.append("National Team")

        # 国家名字
        # 在卡片 div 中查找所有的 div 标签
        all_divs = it.find_all('div')
        # 获取最后一个 div 标签的内容
        last_div = all_divs[-1] if all_divs else None
        # 提取国家名称
        if last_div:
            CountryID=last_div['data-cy']
            Country = last_div.text
            # print(f"Country: {Country}")
            data.append(CountryID)  # 国家代码
            data.append(Country)  # 国家

        datalist.append(data)

    return datalist


# 爬取网页
def getData(baseurl):
    datalist = []  # 用来存储爬取的网页信息
    html = askURL(baseurl)  # 保存获取到的网页源码
    # 2.逐一解析数据
    soup = BeautifulSoup(html, "html.parser")

    # 定义正则表达式，匹配 "" 和 "award-row-NUMBER"
    pattern = re.compile(r"award-row-\d+")

    for ite in soup.find_all('section', {'data-row-id': pattern}):
        item = str(ite)
        # print(item)
        findTitle = ite.find('h2')
        Title = findTitle.text  # 通过正则表达式查找
        # print(str(Title))

        # 项目是单人运动
        if ite.find_all('div', {'data-cy': "single-athlete-award-card"}):
            datalist = get_single_athlete(ite.find_all('div', {'data-cy': "single-athlete-award-card"}), Title,
                                          datalist)
        # 项目是双人运动
        elif ite.find_all('div', {'data-cy': "two-athletes-award-card"}):
            datalist = get_two_athletes(ite.find_all('div', {'data-cy': "two-athletes-award-card"}), Title, datalist)
        # 项目是国家团体
        elif ite.find_all('div', {'data-cy': "team-award-card"}):
            datalist = get_nation_team(ite.find_all('div', {'data-cy': "team-award-card"}), Title, datalist)

    return datalist


# 得到指定一个URL的网页内容
def askURL(url):
    head = {  # 模拟浏览器头部信息，向服务器发送消息
        "User-Agent": "Mozilla / 5.0(Windows NT 10.0; Win64; x64) AppleWebKit / 537.36(KHTML, like Gecko) Chrome / 80.0.3987.122  Safari / 537.36"
    }
    # 用户代理，表示告诉服务器，我们是什么类型的机器、浏览器（本质上是告诉浏览器，我们可以接收什么水平的文件内容）

    request = urllib.request.Request(url, headers=head)
    html = ""
    try:
        response = urllib.request.urlopen(request)
        html = response.read().decode("utf-8")
    except urllib.error.URLError as e:
        if hasattr(e, "code"):
            print(e.code)
        if hasattr(e, "reason"):
            print(e.reason)
    return html


# 保存数据到表格
# def saveData(datalist, savepath, sport):
#     print("save.......")
#     book = openpyxl.Workbook()  # 创建workbook对象
#     sheet = book.active
#     sheet.title = sport  # 创建工作
#     col = ("项目", "奖牌", "运动员链接", "运动员名字", "国家代码", "国家")
#     for i in range(1, 7):
#         sheet.cell(row=1, column=i).value = col[i - 1]  # 列名
#     i = 1
#     for data in datalist:
#         i = i + 1
#         for j in range(1, 7):
#             sheet.cell(row=i, column=j).value = data[j - 1]  # 数据
#     book.save(savepath)  # 保存


def saveData(datalist, savepath, sport,language='zh'):
    # print("save.......")

    if language == 'zh':
        # 创建列名
        columns = ["运动","项目", "奖牌", "运动员链接", "运动员名字", "国家代码", "国家"]
    else:
        columns = ["Sport","Event", "Medal", "Athlete Link", "Athlete Name", "NOC", "Country"]

    # 创建一个空的 DataFrame
    df = pd.DataFrame(columns=columns)

    # 一行一行地添加数据到 DataFrame，并在第一列添加运动名称
    for data in datalist:
        # 在 data 的前面插入 sport，构成新的行
        row_data = [sport] + data
        # 创建一个 DataFrame 单行
        row_df = pd.DataFrame([row_data], columns=columns)
        # 使用 pd.concat 代替 append
        df = pd.concat([df, row_df], ignore_index=True)


    # 将 DataFrame 写入 Excel 文件，并指定 sheet 名
    df.to_excel(savepath, sheet_name=sport, index=False)




if __name__ == "__main__":  # 当程序执行时
    # 调用函数
    main_zh()
    # main_en() 
    print("爬取完毕！")
